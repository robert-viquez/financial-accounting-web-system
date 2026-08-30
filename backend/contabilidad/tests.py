from decimal import Decimal

from django.test import TestCase

from contabilidad.models import CuentaContable
from contabilidad.models import AsientoContable, PeriodoContable
from contabilidad.services import ContabilidadService
from config.test_utils import AccountingAPITestCase
from django.contrib.auth.models import Group
from django.utils import timezone
from rest_framework import status
from io import BytesIO
from openpyxl import load_workbook


class CuentaContableTests(TestCase):
    def test_ordenamiento_del_catalogo_para_reportes(self):
        CuentaContable.objects.create(
            codigo="2", nombre="Pasivo", tipo="PASIVO", naturaleza="ACREEDORA"
        )
        CuentaContable.objects.create(
            codigo="1", nombre="Activo", tipo="ACTIVO", naturaleza="DEUDORA"
        )
        self.assertEqual(
            list(
                CuentaContable.objects.filter(codigo__in=["1", "2"])
                .values_list("codigo", flat=True)
            ),
            ["1", "2"],
        )


class FlujoContableTests(AccountingAPITestCase):
    def setUp(self):
        super().setUp()
        group, _ = Group.objects.get_or_create(name="Contabilidad")
        self.user.groups.add(group)

    def test_venta_genera_asiento_balanceado_y_reportes_reales(self):
        response = self.client.post(
            "/api/ventas/",
            self.venta_payload(tipo="CREDITO"),
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        asiento = AsientoContable.objects.get(origen="VENTA")
        self.assertEqual(asiento.estado, "CONTABILIZADO")
        self.assertEqual(asiento.total_debe, asiento.total_haber)
        self.assertEqual(asiento.detalles.count(), 4)

        diario = self.client.get("/api/reportes-contables/libro-diario/")
        balance = self.client.get("/api/reportes-contables/balance-comprobacion/")
        resultados = self.client.get("/api/reportes-contables/estado-resultados/")
        self.assertEqual(diario.status_code, status.HTTP_200_OK)
        self.assertEqual(len(diario.data), 1)
        self.assertEqual(balance.status_code, status.HTTP_200_OK)
        self.assertTrue(balance.data)
        self.assertEqual(resultados.status_code, status.HTTP_200_OK)
        self.assertEqual(len(resultados.data), 3)

    def test_venta_contado_expone_monto_comercial_y_movimientos_balanceados(self):
        self.producto.precio_venta = Decimal("22050.00")
        self.producto.costo_promedio = Decimal("16367.58")
        self.producto.save(update_fields=["precio_venta", "costo_promedio"])
        response = self.client.post(
            "/api/ventas/",
            self.venta_payload(detalles=[{
                "producto": self.producto.pk,
                "cantidad": "1.00",
                "precio_unitario": "22050.00",
                "descuento": "0.00",
            }]),
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        asiento = AsientoContable.objects.get(origen="VENTA")
        self.assertEqual(AsientoContable.objects.filter(origen="VENTA").count(), 1)
        self.assertEqual(asiento.monto_transaccion, Decimal("22050.00"))
        self.assertEqual(asiento.total_debe, Decimal("38417.58"))
        self.assertEqual(asiento.total_haber, Decimal("38417.58"))

        diario = self.client.get("/api/reportes-contables/libro-diario/")
        data = diario.data[0]
        self.assertEqual(Decimal(data["monto_transaccion"]), Decimal("22050.00"))
        movimientos = {item["tipo"]: item for item in data["movimientos"]}
        self.assertEqual(set(movimientos), {"VENTA", "COSTO_VENTA"})
        self.assertEqual(movimientos["VENTA"]["nombre"], "Reconocimiento de la venta")
        self.assertEqual(Decimal(movimientos["VENTA"]["total_debe"]), Decimal("22050.00"))
        self.assertEqual(Decimal(movimientos["VENTA"]["total_haber"]), Decimal("22050.00"))
        self.assertEqual(movimientos["COSTO_VENTA"]["nombre"], "Costo de venta / salida de inventario")
        self.assertEqual(Decimal(movimientos["COSTO_VENTA"]["total_debe"]), Decimal("16367.58"))
        self.assertEqual(Decimal(movimientos["COSTO_VENTA"]["total_haber"]), Decimal("16367.58"))
        self.assertEqual(
            {linea["cuenta_codigo"] for linea in movimientos["VENTA"]["detalles"]},
            {"1101", "4101"},
        )
        self.assertEqual(
            {linea["cuenta_codigo"] for linea in movimientos["COSTO_VENTA"]["detalles"]},
            {"5101", "1201"},
        )

        ContabilidadService.contabilizar_venta(response.wsgi_request.user.venta_set.get())
        self.assertEqual(AsientoContable.objects.filter(origen="VENTA").count(), 1)
        ContabilidadService.anular_por_origen("VENTA", asiento.referencia)
        asiento.refresh_from_db()
        self.assertEqual(asiento.estado, "ANULADO")
        self.assertEqual(asiento.detalles.count(), 4)

    def test_venta_credito_agrupa_reconocimiento_con_cuentas_por_cobrar(self):
        response = self.client.post(
            "/api/ventas/", self.venta_payload(tipo="CREDITO"), format="json"
        )
        self.assertEqual(response.status_code, status.HTTP_201_CREATED, response.data)
        asiento = AsientoContable.objects.get(origen="VENTA")
        reconocimiento = asiento.detalles.filter(tipo_movimiento="VENTA")
        self.assertEqual(set(reconocimiento.values_list("cuenta__codigo", flat=True)), {"1102", "4101"})
        self.assertEqual(reconocimiento.count(), 2)
        self.assertEqual(asiento.detalles.filter(tipo_movimiento="COSTO_VENTA").count(), 2)

    def test_asiento_manual_desbalanceado_es_rechazado(self):
        cuentas = list(CuentaContable.objects.filter(codigo__in=["1101", "4101"]))
        response = self.client.post(
            "/api/asientos-contables/",
            {
                "fecha": timezone.localdate().isoformat(),
                "descripcion": "Prueba desbalanceada",
                "detalles": [
                    {"cuenta": cuentas[0].pk, "debe": "100.00", "haber": "0.00"},
                    {"cuenta": cuentas[1].pk, "debe": "0.00", "haber": "90.00"},
                ],
            },
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertFalse(AsientoContable.objects.exists())

    def test_asiento_manual_se_contabiliza_sin_duplicarse(self):
        caja = CuentaContable.objects.get(codigo="1101")
        ventas = CuentaContable.objects.get(codigo="4101")
        created = self.client.post(
            "/api/asientos-contables/",
            {
                "fecha": timezone.localdate().isoformat(),
                "descripcion": "Asiento manual",
                "detalles": [
                    {"cuenta": caja.pk, "debe": "100.00", "haber": "0.00"},
                    {"cuenta": ventas.pk, "debe": "0.00", "haber": "100.00"},
                ],
            },
            format="json",
        )
        self.assertEqual(created.status_code, status.HTTP_201_CREATED, created.data)
        posted = self.client.post(
            f"/api/asientos-contables/{created.data['id']}/contabilizar/"
        )
        self.assertEqual(posted.status_code, status.HTTP_200_OK)
        self.assertEqual(posted.data["estado"], "CONTABILIZADO")
        self.assertEqual(AsientoContable.objects.count(), 1)

    def test_periodo_cerrado_bloquea_nuevos_asientos(self):
        today = timezone.localdate()
        PeriodoContable.objects.create(
            nombre="Periodo cerrado",
            fecha_inicio=today,
            fecha_fin=today,
            cerrado=True,
        )
        response = self.client.post(
            "/api/ventas/",
            self.venta_payload(),
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_400_BAD_REQUEST)

    def test_exportacion_xlsx_es_real_y_admite_varias_hojas(self):
        self.client.post("/api/ventas/", self.venta_payload(), format="json")
        response = self.client.post(
            "/api/reportes/exportar/xlsx/",
            {"reportes": ["ventas", "inventario", "balance_comprobacion"]},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b"PK"))
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        self.assertEqual(
            workbook.sheetnames,
            ["Resumen", "Ventas", "Inventario", "Balance de Comprobación"],
        )
        self.assertIsInstance(workbook["Ventas"]["H7"].value, (int, float))

    def test_exportacion_pdf_es_real_y_valida_periodo(self):
        response = self.client.post(
            "/api/reportes/exportar/pdf/",
            {"reportes": ["cxc", "cxp"]},
            format="json",
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertTrue(response.content.startswith(b"%PDF-"))
        invalid = self.client.post(
            "/api/reportes/exportar/pdf/",
            {"reportes": ["ventas"], "desde": "2026-12-31", "hasta": "2026-01-01"},
            format="json",
        )
        self.assertEqual(invalid.status_code, status.HTTP_400_BAD_REQUEST)

    def test_resumen_contable_busca_filtra_ordena_y_totaliza_resultado_completo(self):
        self.client.post("/api/ventas/", self.venta_payload(), format="json")
        response = self.client.get(
            "/api/resumen-contable/",
            {"search": "Ventas", "tipo": "INGRESO", "ordering": "-saldo", "page_size": 10},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK, response.data)
        self.assertEqual(response.data["count"], 1)
        self.assertEqual(response.data["results"][0]["codigo"], "4101")
        self.assertGreater(Decimal(response.data["totals"]["creditos"]), 0)
        self.assertEqual(
            Decimal(response.data["totals"]["creditos"]),
            Decimal(response.data["results"][0]["creditos"]),
        )

    def test_exportacion_resumen_xlsx_respeta_filtros_y_exporta_numeros(self):
        self.client.post("/api/ventas/", self.venta_payload(), format="json")
        response = self.client.get(
            "/api/resumen-contable/exportar/xlsx/",
            {"search": "Caja", "tipo": "ACTIVO"},
        )
        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertIn("resumen_contable_", response["Content-Disposition"])
        workbook = load_workbook(BytesIO(response.content), data_only=True)
        sheet = workbook["Resumen Contable"]
        self.assertEqual(sheet["A7"].value, "1101")
        self.assertIsInstance(sheet["E7"].value, (int, float))
        self.assertIsNone(sheet["A8"].value)
