from collections import OrderedDict
from datetime import date, datetime
from decimal import Decimal
from io import BytesIO

from django.db.models import Sum
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_RIGHT
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import PageBreak, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from compras.models import Compra
from finanzas.models import CuentaPorCobrar, CuentaPorPagar
from inventario.models import Producto
from ventas.models import Venta

from .models import AsientoContable
from .services import ContabilidadService


REPORT_TITLES = OrderedDict([
    ("ventas", "Ventas"),
    ("compras", "Compras"),
    ("inventario", "Inventario"),
    ("cxc", "Cuentas por Cobrar - Antigüedad de Saldos"),
    ("cxp", "Cuentas por Pagar - Antigüedad de Saldos"),
    ("balance_general", "Balance General"),
    ("estado_resultados", "Estado de Resultados"),
    ("balance_comprobacion", "Balance de Comprobación"),
    ("libro_diario", "Libro Diario"),
    ("libro_mayor", "Libro Mayor"),
])

CURRENCY = '₡#,##0.00;[Red]-₡#,##0.00'


def build_accounting_summary_xlsx(rows, totals, period_label):
    """Exporta exactamente el resultado filtrado del resumen, con importes numéricos."""
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Resumen Contable"
    sheet.append(["Queso Los Santos S.A."])
    sheet.append(["Resumen Contable"])
    sheet.append([period_label])
    sheet.append([f"Generado: {timezone.localtime():%d/%m/%Y %H:%M}"])
    sheet.append([])
    sheet.append(["Código", "Cuenta", "Tipo", "Naturaleza", "Débitos", "Créditos", "Saldo"])
    for row in rows:
        sheet.append([
            row["codigo"], row["cuenta"], row["tipo"], row["naturaleza"],
            Decimal(row["debitos"]), Decimal(row["creditos"]), Decimal(row["saldo"]),
        ])
    sheet.append([])
    sheet.append(["TOTALES", "", "", "", Decimal(totals["debitos"]), Decimal(totals["creditos"]), Decimal(totals["saldo"])])
    _style_sheet(sheet, 7, len(rows))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _in_period(queryset, field, desde, hasta):
    if desde:
        queryset = queryset.filter(**{f"{field}__gte": desde})
    if hasta:
        queryset = queryset.filter(**{f"{field}__lte": hasta})
    return queryset


def _aging(fecha_vencimiento, saldo, today):
    days = (today - fecha_vencimiento).days
    if days <= 0:
        bucket = "Vigente / no vencida"
    elif days <= 30:
        bucket = "1–30 días"
    elif days <= 60:
        bucket = "31–60 días"
    elif days <= 90:
        bucket = "61–90 días"
    else:
        bucket = "Más de 90 días"
    return max(days, 0), bucket, Decimal(saldo)


def _accounting_rows(desde, hasta):
    rows = list(ContabilidadService.reporte_saldos(desde, hasta))
    for row in rows:
        debit, credit = Decimal(row["debe"]), Decimal(row["haber"])
        row["saldo"] = debit - credit if row["cuenta__naturaleza"] == "DEUDORA" else credit - debit
    return rows


def report_data(key, desde=None, hasta=None):
    today = timezone.localdate()
    if key == "ventas":
        qs = _in_period(Venta.objects.select_related("cliente", "medio_pago").filter(estado="EMITIDA"), "fecha__date", desde, hasta)
        rows = [[v.fecha.date(), v.numero_comprobante, v.cliente.nombre if v.cliente else "Consumidor final", v.tipo_venta, v.estado, v.subtotal, v.impuesto, v.total] for v in qs]
        return ["Fecha", "Comprobante", "Cliente", "Tipo", "Estado", "Subtotal", "Impuesto", "Total"], rows, {"Total ventas": sum((r[-1] for r in rows), Decimal("0"))}
    if key == "compras":
        qs = _in_period(Compra.objects.select_related("proveedor").filter(estado="REGISTRADA"), "fecha__date", desde, hasta)
        rows = [[c.fecha.date(), c.numero_factura, c.proveedor.nombre, c.tipo_compra, c.estado, c.subtotal, c.impuesto, c.total] for c in qs]
        return ["Fecha", "Factura", "Proveedor", "Tipo", "Estado", "Subtotal", "Impuesto", "Total"], rows, {"Total compras": sum((r[-1] for r in rows), Decimal("0"))}
    if key == "inventario":
        qs = Producto.objects.select_related("categoria", "unidad_medida").order_by("codigo")
        rows = [[p.codigo, p.nombre, p.categoria.nombre, p.stock_actual, p.stock_minimo, p.unidad_medida.simbolo if p.unidad_medida else "", p.costo_promedio, p.precio_venta, p.stock_actual * p.costo_promedio, "Stock bajo" if p.stock_actual <= p.stock_minimo else "Normal"] for p in qs]
        return ["Código", "Producto", "Categoría", "Stock", "Mínimo", "Unidad", "Costo promedio", "Precio venta", "Valor inventario", "Control"], rows, {"Valor de inventario": sum((r[8] for r in rows), Decimal("0")), "Productos con stock bajo": sum(1 for r in rows if r[9] == "Stock bajo")}
    if key in {"cxc", "cxp"}:
        is_cxc = key == "cxc"
        model = CuentaPorCobrar if is_cxc else CuentaPorPagar
        party = "cliente" if is_cxc else "proveedor"
        origin = "venta" if is_cxc else "compra"
        number = "numero_comprobante" if is_cxc else "numero_factura"
        qs = _in_period(model.objects.select_related(party, origin).exclude(estado="ANULADA"), "fecha_emision", desde, hasta)
        rows = []
        buckets = OrderedDict((name, Decimal("0")) for name in ["Vigente / no vencida", "1–30 días", "31–60 días", "61–90 días", "Más de 90 días"])
        for account in qs:
            days, bucket, balance = _aging(account.fecha_vencimiento, account.saldo, today)
            paid = account.monto_original - account.saldo
            rows.append([getattr(account, party).nombre, getattr(getattr(account, origin), number), account.fecha_emision, account.fecha_vencimiento, account.monto_original, paid, account.saldo, account.estado, days, bucket])
            buckets[bucket] += balance
        total = sum((r[6] for r in rows), Decimal("0"))
        overdue = sum((r[6] for r in rows if r[8] > 0), Decimal("0"))
        summary = OrderedDict([("Saldo total", total), ("Saldo vencido", overdue), ("Porcentaje vencido", (overdue / total * 100).quantize(Decimal("0.01")) if total else Decimal("0")), *buckets.items()])
        return ["Cliente" if is_cxc else "Proveedor", "Documento", "Emisión", "Vencimiento", "Monto original", "Pagado", "Saldo", "Estado", "Días vencidos", "Antigüedad"], rows, summary

    balances = _accounting_rows(desde, hasta)
    if key in {"libro_mayor", "balance_comprobacion"}:
        rows = [[r["cuenta__codigo"], r["cuenta__nombre"], r["debe"], r["haber"]] for r in balances]
        headers = ["Código", "Cuenta", "Debe", "Haber"]
        if key == "libro_mayor":
            headers.append("Saldo")
            rows = [row + [balances[index]["saldo"]] for index, row in enumerate(rows)]
        return headers, rows, {"Total debe": sum((r[2] for r in rows), Decimal("0")), "Total haber": sum((r[3] for r in rows), Decimal("0"))}
    if key == "libro_diario":
        qs = _in_period(AsientoContable.objects.filter(estado="CONTABILIZADO").prefetch_related("detalles__cuenta"), "fecha", desde, hasta)
        rows = [[entry.fecha, entry.numero, line.cuenta.codigo, line.cuenta.nombre, line.descripcion or entry.descripcion, line.debe, line.haber] for entry in qs for line in entry.detalles.all()]
        return ["Fecha", "Asiento", "Código", "Cuenta", "Descripción", "Debe", "Haber"], rows, {"Total debe": sum((r[-2] for r in rows), Decimal("0")), "Total haber": sum((r[-1] for r in rows), Decimal("0"))}
    if key == "estado_resultados":
        income = sum((r["saldo"] for r in balances if r["cuenta__tipo"] == "INGRESO"), Decimal("0"))
        expenses = sum((r["saldo"] for r in balances if r["cuenta__tipo"] in {"COSTO", "GASTO"}), Decimal("0"))
        rows = [["Ingresos", income], ["Costos y gastos", expenses], ["Utilidad / Pérdida neta", income - expenses]]
        return ["Concepto", "Monto"], rows, {"Resultado neto": income - expenses}
    if key == "balance_general":
        values = {kind: sum((r["saldo"] for r in balances if r["cuenta__tipo"] == kind), Decimal("0")) for kind in ["ACTIVO", "PASIVO", "PATRIMONIO"]}
        result = sum((r["saldo"] for r in balances if r["cuenta__tipo"] == "INGRESO"), Decimal("0")) - sum((r["saldo"] for r in balances if r["cuenta__tipo"] in {"COSTO", "GASTO"}), Decimal("0"))
        rows = [["Activos", values["ACTIVO"]], ["Pasivos", values["PASIVO"]], ["Patrimonio", values["PATRIMONIO"]], ["Resultado del período", result], ["Pasivos + Patrimonio + Resultado", values["PASIVO"] + values["PATRIMONIO"] + result]]
        return ["Concepto", "Monto"], rows, {"Diferencia contable": values["ACTIVO"] - values["PASIVO"] - values["PATRIMONIO"] - result}
    raise ValueError("Reporte no soportado")


def _period(desde, hasta):
    return f"Período: {desde or 'inicio'} al {hasta or 'hoy'}"


def build_xlsx(keys, desde=None, hasta=None):
    workbook = Workbook()
    workbook.remove(workbook.active)
    summary = workbook.create_sheet("Resumen")
    summary.append(["Queso Los Santos S.A."])
    summary.append(["Resumen de reportes seleccionados"])
    summary.append([_period(desde, hasta)])
    summary.append([f"Generado: {timezone.localtime():%d/%m/%Y %H:%M}"])
    summary.append([])
    summary.append(["Reporte", "Indicador", "Valor"])
    summary.freeze_panes = "A7"

    for key in keys:
        headers, rows, metrics = report_data(key, desde, hasta)
        sheet = workbook.create_sheet(REPORT_TITLES[key][:31])
        sheet.append(["Queso Los Santos S.A."])
        sheet.append([REPORT_TITLES[key]])
        sheet.append([_period(desde, hasta)])
        sheet.append([f"Generado: {timezone.localtime():%d/%m/%Y %H:%M}"])
        sheet.append([])
        sheet.append(headers)
        for row in rows:
            sheet.append(row)
        if not rows:
            sheet.append(["Sin registros para el período seleccionado."])
        else:
            sheet.append([])
            sheet.append(["RESUMEN"])
            for metric, value in metrics.items():
                sheet.append([metric, value])
        for metric, value in metrics.items():
            summary.append([REPORT_TITLES[key], metric, value])
        _style_sheet(sheet, len(headers), len(rows))
    _style_sheet(summary, 3, max(summary.max_row - 6, 0))
    output = BytesIO()
    workbook.save(output)
    return output.getvalue()


def _style_sheet(sheet, width, row_count):
    navy, blue = "17365D", "D9EAF7"
    for row in (1, 2):
        sheet.cell(row, 1).font = Font(bold=True, color="FFFFFF", size=14 if row == 1 else 12)
        for col in range(1, max(width, 1) + 1):
            sheet.cell(row, col).fill = PatternFill("solid", fgColor=navy)
    header_row = 6
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2F75B5")
        cell.alignment = Alignment(horizontal="center")
    sheet.freeze_panes = "A7"
    if width:
        sheet.auto_filter.ref = f"A6:{get_column_letter(width)}{max(6, 6 + row_count)}"
    for row in sheet.iter_rows(min_row=7):
        for cell in row:
            if isinstance(cell.value, Decimal):
                cell.number_format = CURRENCY
                cell.alignment = Alignment(horizontal="right")
            elif isinstance(cell.value, (date, datetime)):
                cell.number_format = "dd/mm/yyyy"
    for index in range(1, width + 1):
        values = [str(sheet.cell(row, index).value or "") for row in range(1, min(sheet.max_row, 100) + 1)]
        sheet.column_dimensions[get_column_letter(index)].width = min(max(max(map(len, values), default=8) + 2, 12), 34)
    sheet.sheet_view.showGridLines = False


def build_pdf(keys, desde=None, hasta=None):
    output = BytesIO()
    doc = SimpleDocTemplate(output, pagesize=landscape(A4), rightMargin=12 * mm, leftMargin=12 * mm, topMargin=14 * mm, bottomMargin=14 * mm)
    styles = getSampleStyleSheet()
    title = ParagraphStyle("ReportTitle", parent=styles["Heading1"], alignment=TA_CENTER, textColor=colors.HexColor("#17365D"), fontSize=15)
    normal = ParagraphStyle("ReportNormal", parent=styles["BodyText"], fontSize=8, leading=10)
    story = []
    for index, key in enumerate(keys):
        headers, rows, metrics = report_data(key, desde, hasta)
        if index:
            story.append(PageBreak())
        story.extend([Paragraph("Queso Los Santos S.A.", title), Paragraph(REPORT_TITLES[key], title), Paragraph(_period(desde, hasta), styles["BodyText"]), Spacer(1, 5 * mm)])
        display_rows = rows or [["Sin registros para el período seleccionado."] + [""] * (len(headers) - 1)]
        data = [[Paragraph(str(cell), normal) for cell in headers]] + [[Paragraph(_pdf_value(cell), normal) for cell in row] for row in display_rows]
        table = Table(data, repeatRows=1, hAlign="LEFT")
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#2F75B5")), ("TEXTCOLOR", (0, 0), (-1, 0), colors.white), ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"), ("GRID", (0, 0), (-1, -1), .3, colors.HexColor("#B7C9DA")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F3F7FA")]), ("VALIGN", (0, 0), (-1, -1), "MIDDLE"), ("ALIGN", (-4, 1), (-1, -1), "RIGHT"), ("LEFTPADDING", (0, 0), (-1, -1), 4), ("RIGHTPADDING", (0, 0), (-1, -1), 4)]))
        story.extend([table, Spacer(1, 4 * mm)])
        for metric, value in metrics.items():
            story.append(Paragraph(f"<b>{metric}:</b> {_pdf_value(value)}", styles["BodyText"]))

    def footer(canvas, document):
        canvas.saveState()
        canvas.setFont("Helvetica", 8)
        canvas.drawString(12 * mm, 8 * mm, "Queso Los Santos S.A. — Reportes")
        canvas.drawRightString(landscape(A4)[0] - 12 * mm, 8 * mm, f"Página {document.page}")
        canvas.restoreState()

    doc.build(story, onFirstPage=footer, onLaterPages=footer)
    return output.getvalue()


def _pdf_value(value):
    if isinstance(value, Decimal):
        return f"CRC {value:,.2f}"
    if isinstance(value, (date, datetime)):
        return value.strftime("%d/%m/%Y")
    return str(value if value is not None else "")
